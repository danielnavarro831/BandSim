import os
import openpyxl
import random


# VARS_LOCATION = os.path.join(os.curdir, 'Data', 'Band_Sim_Vars_V1.xlsx')
VARS_LOCATION = 'C:\\Users\\dnavr\\PycharmProjects\\bandSim\\Data\\Band_Sim_Vars_V1.xlsx'
# STRINGS_LOCATION = os.path.join(os.curdir, 'Data', 'Band_Sim_Strings_V1.xlsx')
STRINGS_LOCATION = ""


class DocReader:
    _vars_location = VARS_LOCATION
    _strings_location = STRINGS_LOCATION
    _member_names_in_use = []

    @classmethod
    def get_string(cls, string: str, args: dict):  # string is the name of the string, args is a dict of the words that
        # replace the variables like "BandName"
        message = ""  # initializes a blank message
        book = openpyxl.load_workbook(DocReader._strings_location)  # Opens the excel file
        sheet = book["Strings"]  # gets the tab titled "Strings" in the opened excel file
        for a in range(2, sheet.max_row + 1):  # Goes through each row of strings and tries to find a matching name
            if sheet.cell(row=a, column=1).value == string:
                message = sheet.cell(row=a, column=2).value  # sets "message" to the text in the string
                break
        if message:  # if a match was found, replace the variable words with the words given
            for a in args.keys():
                message = message.replace(a, str(args[a]), message.count(a))
            return message  # return the message to post a response
        message = "String Error - No Reference: " + string + " " + str(args)
        # If no message was found, return the String Error
        return message

    @classmethod
    def get_random_variable(cls, var_type: str):
        variable_text = ""
        var_type_list = DocReader.get_var_type_sheets(var_type)  # ["First Name", "Last Name"] excel page names
        book = openpyxl.load_workbook(DocReader._vars_location)
        plural = False
        dupe = True
        while dupe:
            for i in range(len(var_type_list)):
                page = var_type_list[i]
                sheet = book[page]
                x = random.randint(1, sheet.max_row)
                variable_text += str(sheet.cell(row=x, column=1).value)
                if var_type == "Band" and i == 0:
                    if sheet.cell(row=x, column=2).value:
                        plural = True
                if i < len(var_type_list) - 1:
                    variable_text += " "
            if plural:
                variable_text += "s"
            if var_type != "Member":
                dupe = False
            else:
                if variable_text not in DocReader._member_names_in_use:
                    dupe = False
        return variable_text

    @classmethod
    def add_member_name_in_use(cls, member_name: str):
        DocReader._member_names_in_use.append(member_name)

    @classmethod
    def remove_member_name_in_use(cls, member_name: str):
        DocReader._member_names_in_use.remove(member_name)

    @classmethod
    def get_var_type_sheets(cls, var_type: str):
        if var_type == "Member":
            return ["First Name", "Last Name"]
        elif var_type == "Band":
            return ["Band1", "Band2", "Band3"]
        elif var_type == "Location":
            return ["Location Name", "Location Type"]
        elif var_type == "Instrument":
            return ["Instrument"]
        elif var_type == "Genre":
            return ["Genre"]
        return []

    @classmethod
    def get_all_instruments(cls):
        instruments = []
        book = openpyxl.load_workbook(DocReader._vars_location)  # Opens the excel file
        sheet = book["Instrument"]
        for i in range(sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value:
                instrument = sheet.cell(row=i, column=1).value
                instruments.append(instrument)
        return instruments

    @classmethod
    def get_all_genres(cls):
        genres = []
        book = openpyxl.load_workbook(DocReader._vars_location)  # Opens the excel file
        sheet = book["Genre"]
        for i in range(sheet.max_row + 1):
            if sheet.cell(row=i + 1, column=1).value:
                genre = sheet.cell(row=i + 1, column=1).value
                genres.append(genre)
        return genres

    @classmethod
    def get_album_types(cls):
        albums = {}
        book = openpyxl.load_workbook(DocReader._vars_location)  # Opens the excel file
        sheet = book["Album Type"]
        for i in range(sheet.max_row + 1):
            if sheet.cell(row=i + 1, column=1).value:
                album_type = sheet.cell(row=i + 1, column=1).value
                album_multiplier = sheet.cell(row=i + 1, column=2).value
                albums[album_type] = album_multiplier
        return albums
